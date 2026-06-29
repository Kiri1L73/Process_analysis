import tkinter as tk
from tkinter import ttk, messagebox
import psutil
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

# Пороги для превышения CPU, RAM (%)
CPU_THRESHOLD = 80
RAM_THRESHOLD = 80

class Process_analysis:
    def __init__(self, root):
        self.root = root
        self.root.title("Приложение анализа и защиты персонального компьютера")
        self.root.geometry("900x500")

        # Переменные для фильтрации
        self.filter_var = tk.StringVar()
        self.filter_var.trace('w', self.on_filter_change)

        # Фильтрация по имени
        filter_frame = tk.Frame(root)
        filter_frame.pack(pady=5, fill=tk.X, padx=10)
        tk.Label(filter_frame, text="Фильтрация по имени:").pack(side=tk.LEFT)
        tk.Entry(filter_frame, textvariable=self.filter_var, width=30).pack(side=tk.LEFT, padx=5)
        tk.Button(filter_frame, text="Сбросить фильтрацию", command=self.clear_filter).pack(side=tk.LEFT, padx=5)

        # Таблица запущенных процессов
        columns = ("PID", "Имя", "CPU %", "RAM %", "Путь")
        self.tree = ttk.Treeview(root, columns=columns, show="headings")
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150 if col != "Путь" else 300)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Кнопки управления приложением
        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=5)
        tk.Button(btn_frame, text="Завершить выбранный процесс", command=self.kill_selected).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Обновить", command=self.update_process).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Выход из приложения", command=root.quit).pack(side=tk.LEFT, padx=5)

        # Статусная строка
        self.status_var = tk.StringVar()
        self.status_var.set("Готов")
        status_label = tk.Label(root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W)
        status_label.pack(side=tk.BOTTOM, fill=tk.X)

        # Хранение списка процессов
        self.processes = []
        # Предотвращение уведомлений подряд
        self.notification_cooldown = 30  # 30 секунд между уведомлениями для каждого PID
        self.last_notification_time = {}

        # Инициализация файла с подозрительными процессами
        created = self.create_file()
        if created:
            messagebox.showinfo("Создание файла с перегружающими процессами", "Создан файл <Перегружающие процессы в системе.xlsx>.")

        # Запуск обновлений
        self.update_process()
        self.update_processes()

    def create_file(self): # Функция создания Excel-файла
        if not os.path.exists('Перегружающие процессы в системе.xlsx'):
            wb = Workbook()
            ws = wb.active
            ws.title = "Перегрузки"
            headers = ['Дата и время', 'PID', 'Имя', 'CPU %', 'RAM %', 'Путь']
            ws.append(headers)
            # Установка ширины столбцов по длине заголовков
            for col_idx, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = len(header) + 2
            wb.save('Перегружающие процессы в системе.xlsx')
            return True
        else:
            try:
                wb = load_workbook('Перегружающие процессы в системе.xlsx')
                ws = wb.active
                if ws.max_row == 0 or ws.cell(1, 1).value is None:
                    headers = ['Дата и время', 'PID', 'Имя', 'CPU %', 'RAM %', 'Путь']
                    ws.append(headers)
                    for col_idx, header in enumerate(headers, 1):
                        col_letter = get_column_letter(col_idx)
                        ws.column_dimensions[col_letter].width = len(header) + 2
                    wb.save('Перегружающие процессы в системе.xlsx')
            except Exception:
                os.remove('Перегружающие процессы в системе.xlsx')
                return self.create_file()
            return False

    def load_file(self, proc): # Функция записи информации о перегружающих процессах в Excel-файл
        try:
            wb = load_workbook('Перегружающие процессы в системе.xlsx')
            ws = wb.active

            new_row = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                proc['pid'],
                proc['name'],
                f"{proc['cpu']:.1f}",
                f"{proc['ram']:.1f}",
                proc['exe']
            ]
            ws.append(new_row)

            # Пересчёт максимальной длины для каждого столбца
            headers = ['Дата и время', 'PID', 'Имя', 'CPU %', 'RAM %', 'Путь']
            max_lengths = [len(h) for h in headers]

            for row in ws.iter_rows(values_only=True):
                for col_idx, value in enumerate(row):
                    if value is not None:
                        length = len(str(value))
                        if length > max_lengths[col_idx]:
                            max_lengths[col_idx] = length

            # Установка ширины столбцов (с запасом 2 символа)
            for col_idx, max_len in enumerate(max_lengths, 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max_len + 2

            wb.save('Перегружающие процессы в системе.xlsx')

        except Exception as e:
            self.status_var.set(f"Ошибка записи файла: {e}")

    def get_processes(self): # Функция сбора информации о запущенных процессах
        proc_list = []
        for proc in psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_percent', 'exe']):
            try:
                info = proc.info
                cpu = info['cpu_percent'] or 0.0
                mem = info['memory_percent'] or 0.0
                proc_list.append({
                    'pid': info['pid'],
                    'name': info['name'] or 'Неизвестно',
                    'cpu': round(cpu, 1),
                    'ram': round(mem, 1),
                    'exe': info['exe'] or 'Недоступно'
                })
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                continue
        return proc_list

    def update_process(self): # Функция обновления данных в таблице с учётом фильтра
        self.processes = self.get_processes()
        filter_text = self.filter_var.get().strip().lower()

        if filter_text:
            filtered = [p for p in self.processes if filter_text in p['name'].lower()]
        else:
            filtered = self.processes

        for item in self.tree.get_children():
            self.tree.delete(item)

        for proc in filtered:
            pid = proc['pid']
            cpu = proc['cpu']
            ram = proc['ram']
            tags = ()
            if cpu > CPU_THRESHOLD or ram > RAM_THRESHOLD:
                tags = ('overload',)
            self.tree.insert('', tk.END, values=(
                pid,
                proc['name'],
                f"{cpu:.1f}%",
                f"{ram:.1f}%",
                proc['exe']
            ), tags=tags)

        self.tree.tag_configure('overload', background='#ffcccc')

        self.proverka_processes()
        self.status_var.set(f"Всего процессов: {len(self.processes)}, отображается: {len(filtered)}")

    def proverka_processes(self): # Проверка на перегружающие процессы и вывод уведомления
        current_time = time.time()
        for proc in self.processes:
            pid = proc['pid']
            cpu = proc['cpu']
            ram = proc['ram']
            if cpu > CPU_THRESHOLD or ram > RAM_THRESHOLD:
                last_time = self.last_notification_time.get(pid, 0)
                if current_time - last_time > self.notification_cooldown:
                    msg = f"Процесс {proc['name']} (PID: {pid}) перегружает систему:\n"
                    if cpu > CPU_THRESHOLD:
                        msg += f"  CPU: {cpu:.1f}% (порог {CPU_THRESHOLD}%)\n"
                    if ram > RAM_THRESHOLD:
                        msg += f"  RAM: {ram:.1f}% (порог {RAM_THRESHOLD}%)\n"
                    messagebox.showwarning("Внимание! Перегрузка системы!", msg)
                    self.last_notification_time[pid] = current_time
                    self.load_file(proc)

    def update_processes(self): # Автоматическое обновление через 2 секунды
        self.update_process()
        self.root.after(2000, self.update_processes)

    def on_filter_change(self, *args): # Обработка изменения фильтра → обновление списка
        self.update_process()

    def clear_filter(self): # Сброс фильтрации по имени
        self.filter_var.set("")
        self.update_process()

    def kill_selected(self): # Завершение выбранного процесса
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Процесс не выбран!", "Пожалуйста, выберите процесс в таблице!")
            return

        item = selection[0]
        try:
            pid = int(self.tree.item(item, 'values')[0])
        except ValueError:
            messagebox.showerror("Ошибка!", "Некорректный PID!")
            return

        if messagebox.askyesno("Подтверждение!", f"Вы действительно хотите завершить процесс с PID {pid}?"):
            try:
                proc = psutil.Process(pid)
                proc.terminate()
                self.status_var.set(f"Процесс {pid} завершён!")
                messagebox.showinfo("Успешное завершение процесса!", f"Процесс с PID {pid} успешно завершён!")
                self.update_process()
            except psutil.NoSuchProcess:
                messagebox.showerror("Ошибка!", f"Процесс с PID {pid} уже не существует!")
            except psutil.AccessDenied:
                messagebox.showerror("Ошибка!", f"Недостаточно прав для завершения процесса {pid}!")
            except Exception as e:
                messagebox.showerror("Ошибка!", f"Не удалось завершить процесс: {e}!")

def main():
    root = tk.Tk()
    app = Process_analysis(root)
    root.mainloop()

if __name__ == "__main__":
    main()